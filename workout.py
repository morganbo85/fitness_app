# 8/18/2023
# Bo Morgan

import tkinter as tk
from tkinter import messagebox, filedialog
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import os.path

# Check if the Excel file exists, if not, create a new one
excel_filename = "C:/test/workout_log.xlsx" # Will need to change path
if not os.path.exists(excel_filename):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Workout Log"
    headers = ["Workout Name", "Sets", "Reps", "Weight"]
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num, value=header)
        cell.alignment = Alignment(horizontal="center")
    workbook.save(excel_filename)
else:
    workbook = openpyxl.load_workbook(excel_filename)
    worksheet = workbook.active

# Add headers to the worksheet
headers = ["Workout Name", "Sets", "Reps", "Weight"]
for col_num, header in enumerate(headers, 1):
    cell = worksheet.cell(row=1, column=col_num, value=header)
    cell.alignment = Alignment(horizontal="center")

# Function to handle the "Save" button click
def save_entry():
    workout_name = entry_workout_name.get()
    sets = entry_sets.get()
    reps = entry_reps.get()
    weight = entry_weight.get()

    if workout_name and sets and reps and weight:
        row = [workout_name, sets, reps, weight]
        worksheet.append(row)
        workbook.save(excel_filename)
        messagebox.showinfo("Success", "Entry saved successfully!")
    else:
        messagebox.showwarning("Incomplete Data", "Please fill in all fields.")

# Create the GUI window
root = tk.Tk()
root.title("Workout Log")

w = 400 # width for the Tk root
h = 350 # height for the Tk root

# get screen width and height
ws = root.winfo_screenwidth() # width of the screen
hs = root.winfo_screenheight() # height of the screen

# calculate x and y coordinates for the Tk root window
x = (ws/2) - (w/2)
y = (hs/2) - (h/2)

# set the dimensions of the screen 
# and where it is placed
root.geometry('%dx%d+%d+%d' % (w, h, x, y))


# Create a frame to hold labels, entry fields, and button
frame = tk.Frame(root)
frame.pack(padx=20, pady=20, expand=True)

# Create labels and entry fields
label_workout_name = tk.Label(root, text="Workout Name:")
entry_workout_name = tk.Entry(root)

label_sets = tk.Label(root, text="Sets:")
entry_sets = tk.Entry(root)

label_reps = tk.Label(root, text="Reps:")
entry_reps = tk.Entry(root)

label_weight = tk.Label(root, text="Weight:")
entry_weight = tk.Entry(root)

# Create and configure the "Save" button
save_button = tk.Button(root, text="Save Entry", command=save_entry)

# Pack the labels, entry fields, and button into the frame
label_workout_name.pack(fill="both", padx=10, pady=5)
entry_workout_name.pack(fill="both", padx=10, pady=5)

label_sets.pack(fill="both", padx=10, pady=5)
entry_sets.pack(fill="both", padx=10, pady=5)

label_reps.pack(fill="both", padx=10, pady=5)
entry_reps.pack(fill="both", padx=10, pady=5)

label_weight.pack(fill="both", padx=10, pady=5)
entry_weight.pack(fill="both", padx=10, pady=5)

save_button.pack(fill="both", padx=10, pady=20)

root.mainloop()
